import pandas as pd
import numpy as np
from sklearn.model_selection import GridSearchCV, cross_val_score, RepeatedKFold, cross_val_predict
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from openpyxl import load_workbook

# Load the dataset
gg = 4  # Replace with your group number
df = pd.read_csv('ds%02d_alt.csv' % (gg))

# Select features and target variable
features = df.iloc[:, :-1]
target = df.iloc[:, -1]

# Normalize the dataset using Min-Max scaling for both cases
scaler = MinMaxScaler()
X_A = scaler.fit_transform(features)

# Case B: Best two features according to PCA
pca = PCA(n_components=2)
features_B = pca.fit_transform(features)
scaler_B = MinMaxScaler()
X_B = scaler_B.fit_transform(features_B)

# Parameters for grid search
param_grid = {
    'n_estimators': [20, 40, 60],
    'min_samples_leaf': [5, 10],
    'criterion': ['gini', 'entropy']
}

# Number of repetitions and folds
num_repetitions = 3
num_folds = 5

# Initialize lists to store metric values
accuracy_values_A = []
precision_values_A = []
recall_values_A = []
f1_values_A = []

accuracy_values_B = []
precision_values_B = []
recall_values_B = []
f1_values_B = []

# Loop for repetitions
for i in range(num_repetitions):
    # Case A: All features
    rf_classifier_A = RandomForestClassifier()
    kf_A = RepeatedKFold(n_splits=num_folds, n_repeats=1)
    
    # Perform cross-validation for Case A
    grid_search_A = GridSearchCV(estimator=rf_classifier_A, param_grid=param_grid, cv=kf_A)
    grid_search_A.fit(X_A, target)
    best_rf_A = grid_search_A.best_estimator_
    
    accuracy_values_A.extend(cross_val_score(best_rf_A, X_A, target, cv=kf_A, scoring='accuracy'))
    predictions_A = cross_val_predict(best_rf_A, X_A, target, cv=kf_A)
    precision_values_A.extend(precision_score(target, predictions_A, average=None))
    recall_values_A.extend(recall_score(target, predictions_A, average=None))
    f1_values_A.extend(f1_score(target, predictions_A, average=None))
    
    # Case B: Best two features according to PCA   
    rf_classifier_B = RandomForestClassifier()
    kf_B = RepeatedKFold(n_splits=num_folds, n_repeats=1)
    
    # Perform cross-validation for Case B
    grid_search_B = GridSearchCV(estimator=rf_classifier_B, param_grid=param_grid, cv=kf_B)
    grid_search_B.fit(X_B, target)
    best_rf_B = grid_search_B.best_estimator_
    
    accuracy_values_B.extend(cross_val_score(best_rf_B, X_B, target, cv=kf_B, scoring='accuracy'))
    predictions_B = cross_val_predict(best_rf_B, X_B, target, cv=kf_B)
    precision_values_B.extend(precision_score(target, predictions_B, average=None))
    recall_values_B.extend(recall_score(target, predictions_B, average=None))
    f1_values_B.extend(f1_score(target, predictions_B, average=None))

# Calculate mean and standard deviation for each metric
mean_accuracy_A = np.mean(accuracy_values_A)
std_accuracy_A = np.std(accuracy_values_A)
mean_precision_A = np.mean(precision_values_A)
std_precision_A = np.std(precision_values_A)
mean_recall_A = np.mean(recall_values_A)
std_recall_A = np.std(recall_values_A)
mean_f1_A = np.mean(f1_values_A)
std_f1_A = np.std(f1_values_A)

mean_accuracy_B = np.mean(accuracy_values_B)
std_accuracy_B = np.std(accuracy_values_B)
mean_precision_B = np.mean(precision_values_B)
std_precision_B = np.std(precision_values_B)
mean_recall_B = np.mean(recall_values_B)
std_recall_B = np.std(recall_values_B)
mean_f1_B = np.mean(f1_values_B)
std_f1_B = np.std(f1_values_B)

# Update the Excel file with mean and standard deviation values
wb = load_workbook(filename="results.xlsx")
ws = wb.active

ws["B9"] = mean_accuracy_A
ws["C9"] = std_accuracy_A
ws["D9"] = mean_precision_A
ws["E9"] = std_precision_A
ws["F9"] = mean_recall_A
ws["G9"] = std_recall_A
ws["H9"] = mean_f1_A
ws["I9"] = std_f1_A

ws["J9"] = mean_accuracy_B
ws["K9"] = std_accuracy_B
ws["L9"] = mean_precision_B
ws["M9"] = std_precision_B
ws["N9"] = mean_recall_B
ws["O9"] = std_recall_B
ws["P9"] = mean_f1_B
ws["Q9"] = std_f1_B

# Save the updated Excel file
wb.save("results.xlsx")

'''
Task 7 introduces Random Forest classifiers with grid search for hyperparameter tuning.

In Case A, leveraging all features, the Random Forest model demonstrates remarkable performance across metrics.
With an average accuracy of 0.911 and an F1 score of 0.909, it achieves robust predictive capabilities. 
The precision, recall, and F1 scores' tight standard deviations (0.013, 0.018, and 0.003, respectively) 
signify consistent and reliable performance across cross-validation folds. Compared to previous tasks, 
Task 7's Case A outperforms all previous cases, showcasing the efficacy of Random Forests in handling multiple features

However, Case B, utilizing only the two best PCA-derived features, 
experiences a notable drop in performance compared to Case A. While maintaining acceptable values, 
with an average accuracy of 0.763 and an F1 score of 0.745, there's a clear decrease in predictive 
power and consistency. The larger standard deviations in precision, recall, and F1 score 
(0.031, 0.067, and 0.022, respectively) indicate greater variability in performance across folds. 
This parallels the findings from Task 6, emphasizing the importance of feature selection and its impact on model performance.

The stark contrast in performance between Case A and Case B underscores the significance
of feature selection in model accuracy. It reaffirms that a larger feature set can provide more
information for models to generalize better, while reduced feature sets might sacrifice predictive performance.
These observations align with the findings across previous tasks, emphasizing the critical role of feature engineering
in enhancing model accuracy and reliability.
'''