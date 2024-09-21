import pandas as pd
from sklearn.model_selection import train_test_split, cross_val_score, RepeatedStratifiedKFold
from sklearn.preprocessing import MinMaxScaler
from sklearn.naive_bayes import GaussianNB
from sklearn.decomposition import PCA
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
import numpy as np
from sklearn.model_selection import cross_val_predict
from openpyxl import load_workbook

# Load the dataset
gg = 4  # Replace with your group number
df = pd.read_csv('ds%02d_alt.csv' % (gg))

# Separate features and target
X = df.iloc[:, :-1]  # Features
y = df.iloc[:, -1]   # Target

# PCA for feature reduction (Case B)
pca = PCA(n_components=2)
X_pca = pca.fit_transform(X)

# Normalize the dataset using Min-Max scaling for CASE B
scaler = MinMaxScaler()
X_B = scaler.fit_transform(X_pca)

# Normalize the dataset using Min-Max scaling for CASE A
scaler = MinMaxScaler()
X_A = scaler.fit_transform(X)

# Specify repetitions and folds for cross-validation
repetitions = 3
folds = 5

# Initialize lists to store metrics across repetitions
accuracy_A_list, precision_A_list, recall_A_list, f1_A_list = [], [], [], []
accuracy_B_list, precision_B_list, recall_B_list, f1_B_list = [], [], [], []

# Perform cross-validation for both cases (A and B)
for i in range(repetitions):
    # Case A: Using all features
    nb_classifier_A = GaussianNB()
    cv_A = RepeatedStratifiedKFold(n_splits=folds, n_repeats=1, random_state=i)
    scores_A = cross_val_score(nb_classifier_A, X_A, y, scoring='accuracy', cv=cv_A)
    accuracy_A_list.extend(scores_A)

    # Get other metrics in addition to accuracy for Case A
    predictions_A = cross_val_predict(nb_classifier_A, X_A, y, cv=cv_A)
    precision_A_list.extend(precision_score(y, predictions_A, average=None))
    recall_A_list.extend(recall_score(y, predictions_A, average=None))
    f1_A_list.extend(f1_score(y, predictions_A, average=None))

    # Case B: Using PCA for feature reduction
    nb_classifier_B = GaussianNB()
    cv_B = RepeatedStratifiedKFold(n_splits=folds, n_repeats=1, random_state=i)
    scores_B = cross_val_score(nb_classifier_B, X_B, y, scoring='accuracy', cv=cv_B)
    accuracy_B_list.extend(scores_B)

    # Get other metrics in addition to accuracy for Case B
    predictions_B = cross_val_predict(nb_classifier_B, X_B, y, cv=cv_B)
    precision_B_list.extend(precision_score(y, predictions_B, average=None))
    recall_B_list.extend(recall_score(y, predictions_B, average=None))
    f1_B_list.extend(f1_score(y, predictions_B, average=None))

# Calculate average and standard metrics for both cases
accuracy_A_avg = np.mean(accuracy_A_list)
accuracy_A_std = np.std(accuracy_A_list)
precision_A_avg = np.mean(precision_A_list)
precision_A_std = np.std(precision_A_list)
recall_A_avg = np.mean(recall_A_list)
recall_A_std = np.std(recall_A_list)
f1_A_avg = np.mean(f1_A_list)
f1_A_std = np.std(f1_A_list)

accuracy_B_avg = np.mean(accuracy_B_list)
accuracy_B_std = np.std(accuracy_B_list)
precision_B_avg = np.mean(precision_B_list)
precision_B_std = np.std(precision_B_list)
recall_B_avg = np.mean(recall_B_list)
recall_B_std = np.std(recall_B_list)
f1_B_avg = np.mean(f1_B_list)
f1_B_std = np.std(f1_B_list)

# Load the existing Excel file
wb = load_workbook(filename="results.xlsx")
ws = wb.active

# Define the cell positions for writing the results
cell_positions = {
    "B": accuracy_A_avg,
    "C": accuracy_A_std,
    "D": precision_A_avg,
    "E": precision_A_std,
    "F": recall_A_avg,
    "G": recall_A_std,
    "H": f1_A_avg,
    "I": f1_A_std,
    "J": accuracy_B_avg,
    "K": accuracy_B_std,
    "L": precision_B_avg,
    "M": precision_B_std,
    "N": recall_B_avg,
    "O": recall_B_std,
    "P": f1_B_avg,
    "Q": f1_B_std,
}

# Write the results to the specified cells
for col, value in cell_positions.items():
    ws[col + "6"] = value

# Save the updated Excel file
wb.save("results.xlsx")

'''
Case A demonstrated robustness in predicting target classes, 
with an average accuracy of 0.770 and a standard deviation of 0.029, 
suggesting consistent and reliable predictions across repetitions. 
Precision averaged at 0.806 with a standard deviation of 0.103, implying a
high proportion of correctly identified positive instances. 
The recall score at 0.768 with a standard deviation of 0.175 showcased the model's 
ability to capture most positive instances, while the F1 score, 
harmonizing precision and recall, averaged 0.762 with a standard deviation of 0.044, 
indicating a balanced performance in classification.

In contrast, Case B utilizing only the top two PCA-derived features displayed 
a slightly lower average accuracy of 0.646 with a standard deviation of 0.029. 
This suggests a reduction in overall predictive power compared to Case A. 
The precision for Case B averaged at 0.672 with a standard deviation of 0.064, 
indicating a moderate proportion of correctly identified positive instances. 
The recall score, averaging at 0.644 with a standard deviation of 0.201,
showcased the model's ability to capture positive instances, albeit less effectively. 
The F1 score, at 0.630 with a standard deviation of 0.077, indicated a trade-off between precision and recall.

This evaluation highlights the trade-off between model complexity and performance, 
showcasing that utilizing all features (Case A) resulted in a more reliable and balanced 
model in terms of classification accuracy, precision, recall, and F1 score compared to 
reducing features through PCA (Case B). It emphasizes the importance of feature selection 
strategies and their direct impact on model performance and predictive power.
'''