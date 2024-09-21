import pandas as pd
import numpy as np
from sklearn.model_selection import GridSearchCV, cross_val_score, RepeatedKFold, cross_val_predict
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split

# Load the dataset
gg = 4  # Replace with your group number
df = pd.read_csv('ds%02d_alt.csv' % (gg))

# Select features and target variable
features = df.iloc[:, :-1]
target = df.iloc[:, -1]

# Case B: Best two features according to PCA
pca = PCA(n_components=2)
features_B = pca.fit_transform(features)
scaler_B = MinMaxScaler()  # Create a separate scaler for Case B
X_B = scaler_B.fit_transform(features_B)

# Parameters for grid search
param_grid = {
    'n_estimators': [20, 40, 60],
    'min_samples_leaf': [5, 10],
    'criterion': ['gini', 'entropy']
}

# Number of repetitions and folds
num_repetitions = 3
num_folds = 5

accuracy_values_B = []
precision_values_B = []
recall_values_B = []
f1_values_B = []

# Loop for repetitions
for i in range(num_repetitions):
    # Case A: All features
    # Case B: Best two features according to PCA   
    rf_classifier_B = RandomForestClassifier()
    kf_B = RepeatedKFold(n_splits=num_folds, n_repeats=1)
    
    # Perform cross-validation for Case B
    grid_search_B = GridSearchCV(estimator=rf_classifier_B, param_grid=param_grid, cv=kf_B)
    grid_search_B.fit(X_B, target)
    best_rf_B = grid_search_B.best_estimator_
    
    accuracy_values_B.extend(cross_val_score(best_rf_B, X_B, target, cv=kf_B, scoring='accuracy'))
    predictions_B = cross_val_predict(best_rf_B, X_B, target, cv=kf_B)
    precision_values_B.extend(precision_score(target, predictions_B, average=None))
    recall_values_B.extend(recall_score(target, predictions_B, average=None))
    f1_values_B.extend(f1_score(target, predictions_B, average=None))

X_train_B, X_test_B, y_train_B, y_test_B = train_test_split(X_B, target, test_size=0.3, random_state=42)
Xte = X_test_B
yte = y_test_B

def plot_class(c, X, y):
    m1 = ['k', 'w']
    m2 = ['x', 'o']
    i = np.where(y == c)[0]
    plt.scatter(X[i, 0], X[i, 1], c=m1[c], marker=m2[c], label='class %d' % (c))

x1lim = [Xte[:, 0].min(), Xte[:, 0].max()]
x2lim = [Xte[:, 1].min(), Xte[:, 1].max()]

npts = 100
x1s = np.linspace(x1lim[0], x1lim[1], npts)
x2s = np.linspace(x2lim[0], x2lim[1], npts)

m = np.zeros((npts, npts))
for k1, x1 in enumerate(x1s):
    for k2, x2 in enumerate(x2s):
        x = np.array([x1, x2])
        m[k1, k2] = best_rf_B.predict([x])

plt.figure()
plt.imshow(m.T, cmap='RdYlGn', origin='lower', extent=(x1lim[0], x1lim[1], x2lim[0], x2lim[1]))
for c in range(len(np.unique(yte))):
    plot_class(c, Xte, yte)
plt.xlabel('$X_1$')
plt.ylabel('$X_2$')
plt.legend()
plt.show()

# Load the dataset and perform PCA transformation for X_samples
gg_samples = 4  # Replace with your group number
df_samples = pd.read_csv('ds%02d_samples_alt.csv' % (gg_samples))

# Select features for X_samples
features_samples = df_samples.iloc[:, :-1]

# Apply PCA transformation on features_samples using the same PCA object used for the training data
features_B_samples = pca.transform(features_samples)
X_samples = scaler_B.transform(features_B_samples)

# Classify the samples and get the posterior probability for each class
predictions_samples = best_rf_B.predict(X_samples)
probabilities_samples = best_rf_B.predict_proba(X_samples)

# Normalize probabilities
normalized_probabilities_samples = probabilities_samples / np.sum(probabilities_samples, axis=1)[:, np.newaxis]

# Display the predictions and normalized probabilities for X_samples
for i, prediction_sample in enumerate(predictions_samples):
    print(f"Sample {i+1}: Predicted class - {prediction_sample}, Normalized probabilities - {normalized_probabilities_samples[i]}")

# Update the Excel file with predictions and probabilities for each sample
wb = load_workbook(filename="results.xlsx")
ws = wb.active

# Write the probabilities and predicted classes for each sample
for i, (prediction, probabilities) in enumerate(zip(predictions_samples, normalized_probabilities_samples), start=14):
    ws[f'D{i}'] = prediction  # Write the predicted class in column D
    
    # Write the normalized probabilities for class 1 (w1) in column B
    ws[f'B{i}'] = probabilities[0]  # Probability for class w1
    
    # Write the normalized probabilities for class 2 (w2) in column C
    ws[f'C{i}'] = probabilities[1]  # Probability for class w2

# Save the updated Excel file
wb.save("results.xlsx")
