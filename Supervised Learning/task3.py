import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.neighbors import KernelDensity
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook
import numpy as np
from sklearn.decomposition import PCA

# Load the dataset
gg = 4  # Replace this with your group number
df = pd.read_csv('ds%02d_alt.csv' % (gg))

# Separate features and target
features = df.iloc[:, :-1]
target = df.iloc[:, -1]

'''
CASE A: ALL FEATURES
'''
# Normalize the dataset using Min-Max scaling for both cases
scaler = MinMaxScaler()
X_A = scaler.fit_transform(features)

# Splitting the dataset into train and test sets for Case A (all features)
X_train_A, X_test_A, y_train_A, y_test_A = train_test_split(X_A, target, test_size=0.3, random_state=1)

# Case A: Using all features
# Use Kernel Density Estimator with Gaussian kernel
h1 = 1
N = X_train_A.shape[0]
h = h1 / np.sqrt(N)

# Fit Kernel Density Estimator
kde_A = KernelDensity(bandwidth=h, kernel='gaussian')
kde_A.fit(X_train_A)

# Calculate log-likelihood for each sample in the test set for each class
log_likelihoods_A = kde_A.score_samples(X_test_A)

# Predict using log-likelihoods (assuming binary classification)
y_pred_A = (log_likelihoods_A >= 0).astype(int)

# Calculate metrics for Case A
accuracy_A = accuracy_score(y_test_A, y_pred_A)
precision_A = precision_score(y_test_A, y_pred_A)
recall_A = recall_score(y_test_A, y_pred_A)
f1_A = f1_score(y_test_A, y_pred_A)

'''
CASE B: BEST 2 FEATURES ACCORDING TO PCA
'''

# Case B: Using PCA for feature reduction and minmax scaler
pca = PCA(n_components=2)
pca_features = pca.fit_transform(features)
scaler = MinMaxScaler()
X_B = scaler.fit_transform(pca_features)

# Splitting the PCA-transformed features into train and test sets for Case B
X_train_B, X_test_B, y_train_B, y_test_B = train_test_split(X_B, target, test_size=0.3, random_state=1)

# Estimate bandwidth for Gaussian kernel for PCA-transformed features
N_pca = X_train_B.shape[0]
h_pca = h1 / np.sqrt(N_pca)

# Fit Kernel Density Estimator
kde_B = KernelDensity(bandwidth=h_pca, kernel='gaussian')
kde_B.fit(X_train_B)

# Calculate log-likelihood for each sample in the test set for each class
log_likelihoods_B = kde_B.score_samples(X_test_B)

# Predict using log-likelihoods for Case B (assuming binary classification)
y_pred_B = (log_likelihoods_B >= 0).astype(int)

# Calculate metrics for Case B
accuracy_B = accuracy_score(y_test_B, y_pred_B)
precision_B = precision_score(y_test_B, y_pred_B)
recall_B = recall_score(y_test_B, y_pred_B)
f1_B = f1_score(y_test_B, y_pred_B)

# Fill in the results in the Excel file
wb = load_workbook(filename="results.xlsx")
ws = wb.active
ws["B5"] = accuracy_A
ws["D5"] = precision_A
ws["F5"] = recall_A
ws["H5"] = f1_A

ws["J5"] = accuracy_B
ws["L5"] = precision_B
ws["N5"] = recall_B
ws["P5"] = f1_B

wb.save("results.xlsx")

'''
In Task 3, Case A utilizing all features achieved impressive results across various metrics. 
It obtained an accuracy of 0.887, indicating strong predictive ability, coupled with a precision of 0.861 
and a recall of 0.946, demonstrating the model's reliability in identifying positive instances. 
The F1-score harmonized these metrics at 0.901, signifying a robust overall performance.

However, Case B, using only the top two features identified by PCA, showed a decrease in performance. 
Despite maintaining a reasonable accuracy of 0.596, precision, recall, and the F1-score 
dropped notably to 0.579, 0.955, and 0.721, respectively. 
This suggests a trade-off between feature reduction simplicity and predictive power.

The disparity in performance indicates that relying solely on PCA-derived features
might sacrifice overall model accuracy and the balance between precision and recall. 
It emphasizes the importance of thoughtful feature selection to maintain a well-balanced predictive model.
'''