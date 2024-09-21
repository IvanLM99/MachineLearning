import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from openpyxl import load_workbook

# Load the dataset (replace gg with your group number)
gg = 4
df = pd.read_csv('ds%02d_alt.csv' % gg)

# Extracting features and target
X = df.drop(columns=['class'])  # Features
y = df['class']  # Target variable

# Perform PCA to find the two most significant features
pca = PCA(n_components=2)
X_pca = pca.fit_transform(X)

# Normalize 2 features using MinMaxScaler
scaler = MinMaxScaler()
X_B = scaler.fit_transform(X_pca)

# Split the dataset into train and test sets (70-30 split)
X_train, X_test, y_train, y_test = train_test_split(X_B, y, test_size=0.3, random_state=100)

# Initialize and fit the Linear Discriminant Analysis model
lda = LinearDiscriminantAnalysis()
lda.fit(X_train, y_train)

# Separate the training set into classes
X_train_class_1 = X_train[y_train == 0]
X_train_class_2 = X_train[y_train == 1]

# Predict on the test set
y_pred = lda.predict(X_test)

# Calculate performance metrics
accuracy = accuracy_score(y_test, y_pred)
precision = precision_score(y_test, y_pred)
recall = recall_score(y_test, y_pred)
f1 = f1_score(y_test, y_pred)

# Fill in the results in the Excel file
wb = load_workbook(filename="results.xlsx")
ws = wb.active
ws["J4"] = accuracy
ws["L4"] = precision
ws["N4"] = recall
ws["P4"] = f1
wb.save("results.xlsx")

# Calculate mean for each class
mean_class_1 = np.mean(X_train_class_1, axis=0)
mean_class_2 = np.mean(X_train_class_2, axis=0)

# Calculate covariance matrix (same for both classes)
cov_matrix = np.cov(X_train.T)

# Calculate inverse of the covariance matrix
cov_inv = np.linalg.inv(cov_matrix)

# Calculate coefficients for class 1
a_1 = -2 * cov_matrix[0, 0] * (mean_class_1[0] - mean_class_2[0])
b_1 = -2 * cov_matrix[0, 1] * (mean_class_1[1] - mean_class_2[1])
c_1 = mean_class_1 @ np.linalg.inv(cov_matrix) @ mean_class_1.T - mean_class_2 @ np.linalg.inv(cov_matrix) @ mean_class_2.T

# Coefficients for class 2 (omega_2)
a_2 = -2 * cov_matrix[0, 0] * (mean_class_2[0] - mean_class_1[0])
b_2 = -2 * cov_matrix[0, 1] * (mean_class_2[1] - mean_class_1[1])
c_2 = mean_class_2 @ np.linalg.inv(cov_matrix) @ mean_class_2.T - mean_class_1 @ np.linalg.inv(cov_matrix) @ mean_class_1.T

# Calculate g_12(x) by subtracting coefficients
a_12 = a_1 - a_2
b_12 = b_1 - b_2
c_12 = c_1 - c_2

# Get coefficients for the decision boundary
coefficients = lda.coef_[0]
# Get the intercept
intercept = lda.intercept_

# Print the functions
print(f"g_1(x) = {a_1} * x1 + {b_1} * x2 + {c_1}")
print(f"g_2(x) = {a_2} * x1 + {b_2} * x2 + {c_2}")
print(f"g_1_2(x) = {a_12} * x1 + {b_12} * x2 + {c_12}")

# Get coefficients for the decision boundary
coefficients = -lda.coef_[0]
# Get the intercept
intercept = -lda.intercept_

# Plotting the decision boundary
x_min, x_max = X_train[:, 0].min(), X_train[:, 0].max()
y_min, y_max = X_train[:, 1].min(), X_train[:, 1].max()
xx, yy = np.meshgrid(np.arange(x_min, x_max, 0.01), np.arange(y_min, y_max, 0.01))

Z = lda.predict(np.c_[xx.ravel(), yy.ravel()])
Z = Z.reshape(xx.shape)

plt.contourf(xx, yy, Z, alpha=0.4)
plt.scatter(X_train[:, 0], X_train[:, 1], c=y_train, edgecolors='k')
plt.title('Decision Boundary')
plt.xlabel('Feature 1')
plt.ylabel('Feature 2')
plt.show()


'''
Task 2 introduced Linear Discriminant Analysis (LDA) applied to a dataset, 
revealing a model with moderate but noteworthy performance. 
The model achieved an average accuracy of 0.665, signifying a reasonably high rate of correct predictions. 
Precision and recall, averaging at 0.652 and 0.728, respectively, along with an F1-score of 0.688, 
highlighted the model's capability in correctly identifying positive instances while capturing a significant portion of actual positives.

The quadratic equations (g_1(x), g_2(x), and g_1_2(x)) derived from the model's coefficients 
elucidate the decision boundaries between distinct classes. 
While the model demonstrated competent discriminative ability between classes, 
these metrics suggest that there's potential for further enhancement, especially 
in optimizing precision and recall to handle more complex datasets.

The LDA model's performance appears satisfactory, 
especially in its ability to classify data and differentiate between classes. 
However, fine-tuning the model to strike a better balance between precision 
and recall might be beneficial for handling more intricate and nuanced datasets, 
ensuring robust performance in varied scenarios.
'''