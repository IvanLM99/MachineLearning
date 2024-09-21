import pandas as pd
from sklearn.model_selection import train_test_split, cross_val_score, RepeatedStratifiedKFold
from sklearn.preprocessing import MinMaxScaler
from sklearn.linear_model import LogisticRegression
from sklearn.decomposition import PCA
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
import numpy as np
from sklearn.model_selection import cross_val_predict
from openpyxl import load_workbook

# Load the dataset
gg = 4  # Replace with your group number
df = pd.read_csv('ds%02d_alt.csv' % (gg))

# Separate features and target
X = df.iloc[:, :-1]  # Features
y = df.iloc[:, -1]   # Target

# PCA for feature reduction (Case B)
pca = PCA(n_components=2)
X_pca = pca.fit_transform(X)

# Normalize the dataset using Min-Max scaling for CASE B
scaler = MinMaxScaler()
X_B = scaler.fit_transform(X_pca)

# Normalize the dataset using Min-Max scaling for CASE A
scaler = MinMaxScaler()
X_A = scaler.fit_transform(X)

# Specify repetitions and folds for cross-validation
repetitions = 3
folds = 5

# Initialize lists to store metrics across repetitions
accuracy_A_list, precision_A_list, recall_A_list, f1_A_list = [], [], [], []
accuracy_B_list, precision_B_list, recall_B_list, f1_B_list = [], [], [], []

# Perform cross-validation for both cases (A and B)
for i in range(repetitions):
    # Case A: Using all features
    lr_classifier_A = LogisticRegression(penalty=None)
    cv_A = RepeatedStratifiedKFold(n_splits=folds, n_repeats=1, random_state=i)
    scores_A = cross_val_score(lr_classifier_A, X_A, y, scoring='accuracy', cv=cv_A)
    accuracy_A_list.extend(scores_A)

    # Get other metrics in addition to accuracy for Case A
    predictions_A = cross_val_predict(lr_classifier_A, X_A, y, cv=cv_A)
    precision_A_list.extend(precision_score(y, predictions_A, average=None))
    recall_A_list.extend(recall_score(y, predictions_A, average=None))
    f1_A_list.extend(f1_score(y, predictions_A, average=None))

    # Case B: Using PCA for feature reduction
    lr_classifier_B = LogisticRegression(penalty=None)
    cv_B = RepeatedStratifiedKFold(n_splits=folds, n_repeats=1, random_state=i)
    scores_B = cross_val_score(lr_classifier_B, X_B, y, scoring='accuracy', cv=cv_B)
    accuracy_B_list.extend(scores_B)

    # Get other metrics in addition to accuracy for Case B
    predictions_B = cross_val_predict(lr_classifier_B, X_B, y, cv=cv_B)
    precision_B_list.extend(precision_score(y, predictions_B, average=None))
    recall_B_list.extend(recall_score(y, predictions_B, average=None))
    f1_B_list.extend(f1_score(y, predictions_B, average=None))

# Calculate average and standard metrics for both cases
accuracy_A_avg = np.mean(accuracy_A_list)
accuracy_A_std = np.std(accuracy_A_list)
precision_A_avg = np.mean(precision_A_list)
precision_A_std = np.std(precision_A_list)
recall_A_avg = np.mean(recall_A_list)
recall_A_std = np.std(recall_A_list)
f1_A_avg = np.mean(f1_A_list)
f1_A_std = np.std(f1_A_list)

accuracy_B_avg = np.mean(accuracy_B_list)
accuracy_B_std = np.std(accuracy_B_list)
precision_B_avg = np.mean(precision_B_list)
precision_B_std = np.std(precision_B_list)
recall_B_avg = np.mean(recall_B_list)
recall_B_std = np.std(recall_B_list)
f1_B_avg = np.mean(f1_B_list)
f1_B_std = np.std(f1_B_list)

# Load the existing Excel file
wb = load_workbook(filename="results.xlsx")
ws = wb.active

# Define the cell positions for writing the results
cell_positions = {
    "B": accuracy_A_avg,
    "C": accuracy_A_std,
    "D": precision_A_avg,
    "E": precision_A_std,
    "F": recall_A_avg,
    "G": recall_A_std,
    "H": f1_A_avg,
    "I": f1_A_std,
    "J": accuracy_B_avg,
    "K": accuracy_B_std,
    "L": precision_B_avg,
    "M": precision_B_std,
    "N": recall_B_avg,
    "O": recall_B_std,
    "P": f1_B_avg,
    "Q": f1_B_std,
}

# Write the results to the specified cells
for col, value in cell_positions.items():
    ws[col + "7"] = value

# Save the updated Excel file
wb.save("results.xlsx")

'''
Case A displayed an improvement in all metrics compared to Task 4. With all features utilized, 
Task 5's Case A achieved higher values in accuracy (0.815 vs. 0.770), precision (0.817 vs. 0.806), 
recall (0.814 vs. 0.768), and F1 score (0.814 vs. 0.762). This upturn showcases the efficacy of 
logistic regression with the full feature set in this task, potentially owing to optimized model tuning or dataset specifics.

Conversely, Task 5's Case B exhibited similar metrics as Task 4's Case B. 
Both cases utilized PCA for feature reduction and logistic regression for classification. 
Despite similar methodologies, Task 5's Case B showed only slight fluctuations in accuracy (0.642 vs. 0.646),
precision (0.643 vs. 0.672), recall (0.641 vs. 0.644), and F1 score (0.641 vs. 0.630) compared to Task 4. 
This outcome suggests consistency in the predictive capability of a reduced-feature logistic regression model between the tasks.

The variance observed across Task 5's cases echoes the importance of exploring different 
feature selection methods and algorithms. While Case A's improvement in Task 5 signifies potential
enhancements in the model or data optimization, the similarity in Case B between tasks underlines
the consistency of feature reduction and logistic regression within this specific scope.
These findings emphasize the iterative nature of model building, where diverse approaches 
impact model performance, and incremental enhancements can refine predictive capabilities
'''