import pandas as pd
from sklearn.model_selection import train_test_split, cross_val_score, RepeatedStratifiedKFold
from sklearn.preprocessing import MinMaxScaler
from sklearn.naive_bayes import GaussianNB
from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.decomposition import PCA
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from sklearn.ensemble import StackingClassifier
import numpy as np
from openpyxl import load_workbook
from sklearn.model_selection import cross_val_predict

# Load the dataset
gg = 4  # Replace with your group number
df = pd.read_csv('ds%02d_alt.csv' % (gg))

# Separate features and target
X = df.iloc[:, :-1]  # Features
y = df.iloc[:, -1]   # Target

# PCA for feature reduction (Case B)
pca = PCA(n_components=2)
X_pca = pca.fit_transform(X)

# Normalize the dataset using Min-Max scaling for CASE B
scaler = MinMaxScaler()
X_B = scaler.fit_transform(X_pca)

# Normalize the dataset using Min-Max scaling for CASE A
scaler = MinMaxScaler()
X_A = scaler.fit_transform(X)

# Define classifiers for the first layer
nb_classifier = GaussianNB()
lr_classifier = LogisticRegression()
dt_classifier = DecisionTreeClassifier(min_samples_leaf=10)

# Define blender (meta-classifier)
blender = DecisionTreeClassifier(max_depth=3)

# Stacking ensemble model
estimators = [('nb', nb_classifier), ('lr', lr_classifier), ('dt', dt_classifier)]
stacking_classifier = StackingClassifier(estimators=estimators, final_estimator=blender)

# Specify repetitions and folds for cross-validation
repetitions = 3
folds = 5

# Initialize lists to store metrics across repetitions
accuracy_A_list, precision_A_list, recall_A_list, f1_A_list = [], [], [], []
accuracy_B_list, precision_B_list, recall_B_list, f1_B_list = [], [], [], []

# Perform cross-validation for both cases (A and B)
for i in range(repetitions):
    # Case A: Using all features
    cv_A = RepeatedStratifiedKFold(n_splits=folds, n_repeats=1, random_state=i)
    scores_A = cross_val_score(stacking_classifier, X_A, y, scoring='accuracy', cv=cv_A)
    accuracy_A_list.extend(scores_A)

    # Get other metrics in addition to accuracy for Case A
    predictions_A = cross_val_predict(stacking_classifier, X_A, y, cv=cv_A)
    precision_A_list.extend(precision_score(y, predictions_A, average=None))
    recall_A_list.extend(recall_score(y, predictions_A, average=None))
    f1_A_list.extend(f1_score(y, predictions_A, average=None))

    # Case B: Using PCA for feature reduction
    cv_B = RepeatedStratifiedKFold(n_splits=folds, n_repeats=1, random_state=i)
    scores_B = cross_val_score(stacking_classifier, X_B, y, scoring='accuracy', cv=cv_B)
    accuracy_B_list.extend(scores_B)

    # Get other metrics in addition to accuracy for Case B
    predictions_B = cross_val_predict(stacking_classifier, X_B, y, cv=cv_B)
    precision_B_list.extend(precision_score(y, predictions_B, average=None))
    recall_B_list.extend(recall_score(y, predictions_B, average=None))
    f1_B_list.extend(f1_score(y, predictions_B, average=None))

# Calculate average and standard metrics for both cases
accuracy_A_avg = np.mean(accuracy_A_list)
accuracy_A_std = np.std(accuracy_A_list)
precision_A_avg = np.mean(precision_A_list)
precision_A_std = np.std(precision_A_list)
recall_A_avg = np.mean(recall_A_list)
recall_A_std = np.std(recall_A_list)
f1_A_avg = np.mean(f1_A_list)
f1_A_std = np.std(f1_A_list)

accuracy_B_avg = np.mean(accuracy_B_list)
accuracy_B_std = np.std(accuracy_B_list)
precision_B_avg = np.mean(precision_B_list)
precision_B_std = np.std(precision_B_list)
recall_B_avg = np.mean(recall_B_list)
recall_B_std = np.std(recall_B_list)
f1_B_avg = np.mean(f1_B_list)
f1_B_std = np.std(f1_B_list)

# Load the existing Excel file
wb = load_workbook(filename="results.xlsx")
ws = wb.active

# Define the cell positions for writing the results
cell_positions = {
    "B": accuracy_A_avg,
    "C": accuracy_A_std,
    "D": precision_A_avg,
    "E": precision_A_std,
    "F": recall_A_avg,
    "G": recall_A_std,
    "H": f1_A_avg,
    "I": f1_A_std,
    "J": accuracy_B_avg,
    "K": accuracy_B_std,
    "L": precision_B_avg,
    "M": precision_B_std,
    "N": recall_B_avg,
    "O": recall_B_std,
    "P": f1_B_avg,
    "Q": f1_B_std,
}

# Write the results to the specified cells
for col, value in cell_positions.items():
    ws[col + "8"] = value

# Save the updated Excel file
wb.save("results.xlsx")


'''
In Task 6, the stacked ensemble model incorporating Naive Bayes, Logistic Regression, 
and Decision Tree classifiers performed consistently well across both cases A and B. 
This ensemble strategy outperformed individual classifiers in terms of accuracy, precision, recall, and F1 score.

Case A, leveraging all features, showcased substantial predictive capability with an 
accuracy of 0.839 and an F1 score of 0.840. This demonstrates an improvement compared 
to Task 4 and Task 5 for Case A, reflecting the efficacy of the ensemble method in harnessing
 the full feature set. The precision and recall metrics, averaging 0.845 and 0.840, respectively,
 highlight its balanced performance in correctly identifying positive instances while minimizing false positives.

Conversely, Case B's use of reduced features via PCA led to a decline in all metrics 
compared to Task 5's Case A, indicating that the ensemble's predictive capacity diminished
 with the reduced feature set. Despite this reduction, Case B in Task 6 still outperformed 
 Task 4 in terms of precision, recall, and F1 score, showcasing the inherent strength of the
 ensemble approach even with fewer features.

This demonstrates the consistency of the stacked ensemble model across diverse datasets 
and its ability to adapt to different feature subsets. The relationship across tasks 
showcases that while Case A continues to perform well across tasks, Case B's performance
 fluctuates based on the dimensionality reduction method employed, reiterating the importance
 of feature selection techniques in ensemble learning
'''